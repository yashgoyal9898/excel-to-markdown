"""
excel_to_md.py
==============
Excel workbook → Markdown converter.

Hybrid approach:
  • pandas   – actual used-range detection (no phantom empty cols / rows)
  • openpyxl – formatting-aware cell reading (%, dates, currency, …)
"""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ═══════════════════════════════════════════════════════════════════════
#  Utility helpers
# ═══════════════════════════════════════════════════════════════════════

def _escape(value) -> str:
    """Pipe characters aur newlines ko Markdown-safe string mein convert karo."""
    if value is None:
        return ""
    text = str(value)
    text = text.replace("|", "\\|")
    text = text.replace("\r\n", "<br>")
    text = text.replace("\n", "<br>")
    return text


def _decimal_places(fmt: str) -> int:
    """Format string mein explicit decimal-place count nikalo."""
    m = re.search(r'\.([0#]+)', fmt)
    return len(m.group(1)) if m else 0


def _positive_fmt(fmt: str) -> str:
    """
    Excel format strings mein ; se alag positive/negative parts hote hain.
    Sirf pehla (positive-value) part return karo.
    """
    return fmt.split(";")[0] if ";" in fmt else fmt


# ═══════════════════════════════════════════════════════════════════════
#  Date formatting
# ═══════════════════════════════════════════════════════════════════════

# Longest tokens pehle — partial replacement se bachne ke liye
_DATE_TOKENS: list[tuple[str, str]] = [
    ("AM/PM", "%p"), ("am/pm", "%p"),
    ("dddd",  "%A"), ("ddd",   "%a"),
    ("yyyy",  "%Y"), ("yy",    "%y"),
    ("mmmm",  "%B"), ("mmm",   "%b"),
    ("hh",    "%H"), ("h",     "%H"),
    ("ss",    "%S"),
    ("mm",    "%m"), ("m",     "%m"),
    ("dd",    "%d"), ("d",     "%d"),
]


def _excel_date_fmt_to_strftime(fmt: str) -> str:
    """
    Excel date-format tokens → Python strftime tokens.

    Null-byte placeholders use karta hai taaki ek token doosre ke andar
    replace na ho (e.g. "d" replacing inside "dd", or "m" inside "%m").
    """
    result = fmt
    placeholders: dict[str, str] = {}

    for idx, (excel_tok, py_tok) in enumerate(_DATE_TOKENS):
        ph = f"\x00{idx}\x00"
        if excel_tok in result:
            result = result.replace(excel_tok, ph)
            placeholders[ph] = py_tok

    for ph, py_tok in placeholders.items():
        result = result.replace(ph, py_tok)

    return result


def _format_date(value: date | datetime, fmt: str) -> str:
    """Date / datetime value ko Excel format ke according string mein format karo."""
    py_fmt = _excel_date_fmt_to_strftime(_positive_fmt(fmt))
    try:
        return value.strftime(py_fmt)
    except Exception:
        # Fallback to ISO-like format
        if isinstance(value, datetime) and (value.hour or value.minute or value.second):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")


# ═══════════════════════════════════════════════════════════════════════
#  Number formatting
# ═══════════════════════════════════════════════════════════════════════

_FMT_ESCAPE_RE = re.compile(r'["\\_*]')          # Excel escape chars
_PREFIX_RE     = re.compile(r'^([^\d#,.0@?+\-\s]+)')  # Currency / unit prefix


def _format_number(value: int | float, fmt: str) -> str:
    """
    Numeric value ko Excel number-format ke according format karo.
    Handles: comma separators, decimal places, currency prefixes (₹, $, €, …).
    """
    fmt = _positive_fmt(fmt)

    if fmt in ("General", "@", ""):
        # Float jo actually integer hai usse bina .0 ke dikhao
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    use_comma = "#,##" in fmt
    dec = _decimal_places(fmt)

    # Leading currency / unit symbol extract karo
    clean = _FMT_ESCAPE_RE.sub("", fmt)
    pm = _PREFIX_RE.match(clean)
    prefix = pm.group(1).strip() if pm else ""

    try:
        if use_comma:
            formatted = f"{value:,.{dec}f}"
        elif "#" in fmt or "0" in fmt:
            formatted = f"{value:.{dec}f}"
        else:
            return str(value)
        return f"{prefix}{formatted}"
    except Exception:
        return str(value)


# ═══════════════════════════════════════════════════════════════════════
#  Cell formatter  (main dispatch)
# ═══════════════════════════════════════════════════════════════════════

def _format_cell(cell) -> str:
    """
    openpyxl Cell → Markdown-safe string.
    cell.number_format dekh ke appropriate formatter dispatch karo.
    """
    value = cell.value
    if value is None:
        return ""

    fmt = str(cell.number_format or "General")

    # ── Percentage ──────────────────────────────────────────────────── #
    if "%" in fmt and isinstance(value, (int, float)):
        dec = _decimal_places(_positive_fmt(fmt))
        return _escape(f"{value * 100:.{dec}f}%")

    # ── Date / Datetime ─────────────────────────────────────────────── #
    if isinstance(value, (datetime, date)):
        return _escape(_format_date(value, fmt))

    # ── Numeric ─────────────────────────────────────────────────────── #
    if isinstance(value, (int, float)):
        return _escape(_format_number(value, fmt))

    # ── Fallback: strings, booleans, etc. ───────────────────────────── #
    return _escape(value)


# ═══════════════════════════════════════════════════════════════════════
#  Data-bounds detection  (pandas)
# ═══════════════════════════════════════════════════════════════════════

def _get_data_bounds(excel_bytes: BytesIO) -> dict[str, tuple[int, int]]:
    """
    pandas se har sheet ka actual used range detect karo.

    Problem:
        openpyxl ka ws.max_row / max_column formatting-only (empty) cells
        ko bhi count karta hai → thousands of empty Markdown columns.

    Solution:
        pandas automatically trailing empty rows/columns ko ignore karta hai.
        Hum pandas ka use sirf bounds detection ke liye karte hain,
        actual cell values openpyxl se padhte hain.

    Returns:
        {sheet_name: (nrows, ncols)}
        nrows / ncols directly openpyxl iter_rows ke max_row / max_col hain.
    """
    excel_bytes.seek(0)
    xl = pd.ExcelFile(excel_bytes, engine="openpyxl")
    bounds: dict[str, tuple[int, int]] = {}

    for sheet_name in xl.sheet_names:
        try:
            # dtype=object → type-conversion errors se bacho
            # header=None  → sab rows data rows hain (koi row skip nahi)
            df = xl.parse(sheet_name, header=None, dtype=object)
        except Exception:
            bounds[sheet_name] = (0, 0)
            continue

        if df.empty:
            bounds[sheet_name] = (0, 0)
            continue

        has_data = df.notna()
        row_mask = has_data.any(axis=1)
        col_mask = has_data.any(axis=0)

        if not row_mask.any():
            bounds[sheet_name] = (0, 0)
            continue

        # pandas 0-based last index + 1  →  openpyxl 1-based max_row / max_col
        last_row = int(row_mask[::-1].idxmax()) + 1
        last_col = int(col_mask[::-1].idxmax()) + 1
        bounds[sheet_name] = (last_row, last_col)

    return bounds


# ═══════════════════════════════════════════════════════════════════════
#  Worksheet → Markdown
# ═══════════════════════════════════════════════════════════════════════

def _worksheet_to_markdown(ws: Worksheet, nrows: int, ncols: int) -> str:
    """
    Single worksheet ko Markdown table string mein convert karo.

    Sirf actual data range (nrows × ncols) iterate hoga.
    Koi phantom empty columns generate nahi honge.
    """
    if nrows == 0 or ncols == 0:
        return "_Empty Sheet_"

    rows = list(ws.iter_rows(
        min_row=1, max_row=nrows,
        min_col=1, max_col=ncols,
    ))

    if not rows:
        return "_Empty Sheet_"

    header_cells = [_format_cell(c) for c in rows[0]]
    lines: list[str] = [
        "| " + " | ".join(header_cells) + " |",
        "| " + " | ".join(["---"] * len(header_cells)) + " |",
    ]

    for row in rows[1:]:
        lines.append("| " + " | ".join(_format_cell(c) for c in row) + " |")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════
#  Public API
# ═══════════════════════════════════════════════════════════════════════

def extract_excel_to_markdown(excel_bytes: BytesIO, workbook_name: str) -> str:
    """
    Excel workbook (.xlsx) ko complete Markdown string mein convert karo.

    Strategy:
      1. pandas  → har sheet ka actual data bounds detect karo
      2. openpyxl → unhi bounds ke andar formatting-aware cell values paro

    Args:
        excel_bytes:    Excel file ka BytesIO object (reusable — seek(0) internally hoga).
        workbook_name:  Output Markdown ka top-level H1 heading.

    Returns:
        Markdown string — har worksheet ek alag H2 section + table.
    """
    # Step 1: pandas se actual data bounds (empty cols/rows automatically excluded)
    bounds = _get_data_bounds(excel_bytes)

    # Step 2: openpyxl se workbook load (data_only=True → formula results milenge)
    excel_bytes.seek(0)
    wb = load_workbook(excel_bytes, data_only=True)

    md: list[str] = [f"# {workbook_name}", ""]

    for ws in wb.worksheets:
        nrows, ncols = bounds.get(ws.title, (0, 0))
        md.extend([
            f"## {ws.title}",
            "",
            _worksheet_to_markdown(ws, nrows, ncols),
            "",
            "---",
            "",
        ])

    return "\n".join(md)
